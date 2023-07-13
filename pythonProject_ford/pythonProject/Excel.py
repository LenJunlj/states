
from openpyxl import workbook
wb = workbook.Workbook()
ws = wb.active
ws.title = "DDM"
for i in range(2,22):
    ws.cell(row=i, column=1, value = 'PASS')
wb.save(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
# for i in range(1,11):
#
# #新建工作簿
#     wb = workbook.Workbook()
#     #新建工作表并修改名称
#     ws = wb.active
#     ws.title = 'N' + str(i)
#     # 写入标题
#     ws['A1'] = 'class'
#     ws['B1'] = 'name'
#     # 循环写入内容
#     for j in range(2,22):
#         ws.cell(row=j, column=1, value = 'N' + str(i))
#         ws.cell(row=j, column=2, value = 'st_' + str(j-1))
#         #生成1-100之间的成绩（整数）
#     wb.save(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject' + str(i) + '.xlsx')
