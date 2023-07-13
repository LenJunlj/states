import paramiko
import time
import sys
import openpyxl


# 创建SSH客户端
ssh_client = paramiko.SSHClient()

# 允许自动添加主机密钥
ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# 导入密钥
private_key_path = 'id_ecg'
private_key = paramiko.RSAKey.from_private_key_file(private_key_path)

# 连接到远程服务器
ssh_client.connect('10.2.0.1', username='root', pkey=private_key)

# 执行命令并将输出重定向到文件
command = 'fdpcat'
stdin, stdout, stderr = ssh_client.exec_command(command)

# 初始化输出列表
output_list = []


# 每2秒读取一次输入内容并打印前2秒内的所有内容
start_time = time.time()
while True:
    if stdout.channel.recv_ready():
        output = stdout.channel.recv(1024).decode('latin-1')
        # print(output)
        output_list.append(output)
    else:
        time.sleep(0.1)
    # Case11 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    case11_flag = "OTAM_S1011"
    # 判断是否达到2秒
    if time.time() - start_time >= 2:
        print("前2秒内的输出内容:")
        # print(time.time())
        # print(start_time)
        for item in output_list:
            print(item)
            # # 将数据输出文件中，注意点1. 所指定的盘存在，2. 使用file=
            # fp = open(r"C:\Users\YYU66\Desktop\Test\Overnight Test\test.fdp", "a+")  # a+ 如果文件不存在就创建。存在就在文件内容的后面继续追加
            # fp.write(item)
            # # fp.close()
            # with open(r"C:\Users\YYU66\Desktop\Test\Overnight Test\test.fdp", "r") as file:
            #     for line in file:
            #         try:
#                         if case11_flag in line:
#                             print(happypath)
#                             wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
#                             ws = wb.active
#                             # 获取最后一行的行号
#                             last_row = ws.max_column + 1
#                             ws.cell(row=last_row, column=2, value=happypath)
#                             wb.save(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
#                             break
#                     except:
#                         print(unhappypath)
#                         wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
#                         ws = wb.active
#                         # 获取最后一行的行号
#                         last_row = ws.max_row + 1
#                         ws.cell(row=last_row, column=2, value=unhappypath)
#                         wb.save('result.xlsx')
#                         break
#
#
#         print("------------------------------------")
        # 重置计时器和输出列表
        start_time = time.time()
        output_list = []
# 读取输出内容

ssh_client.close()



