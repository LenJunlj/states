import paramiko
import time
import openpyxl
from openpyxl import workbook
import canoe_open
from canoe_open import canalyzer

import Excel_text
from pycan import crash
from canoe_open import canalyzer
from Excel_text import *
# class testcase:
def crash120s(case_flag,butt):
    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()
    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)

    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)

    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)

    # 初始化输出列表
    output_list = []


    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    current_time = time.localtime()
    formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", current_time)
    print(formatted_time)

    # # Case11 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 600:
            # fail的应用场景
            print(unhappypath)
            Excel_text.enter_FAIL('B15')
            Excel_text.enter_Time('C15',formatted_time)
            canalyzer.close_cfg()
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)
        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                # print(type(item))
                if case_flag in item:
                    print(happypath)
                    Excel_text.enter_PASS('B15')
                    Excel_text.enter_Time('C15',formatted_time)
                    test_status = 1
                    canalyzer().stop_Measurement()  #关闭CAN工程
                    canalyzer.close_cfg()
                    break
            # print("------------------------------------")
            # 重置计时器和输出列表
            start_time = time.time()
            output_list = []

def crash1023(case_flag,butt):
    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()
    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)
    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)
    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)
    # 初始化输出列表
    output_list = []

    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    search_status = 0
    # Case11 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    case_flag1 = "OTAM_S1011"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 600:
            # fail的应用场景
            print(unhappypath)
            Excel_text.enter_FAIL('B14')
            Excel_text.enter_Time('C14', start_time)
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)

        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                if search_status == 1:
                    if case_flag1 in item:
                        print(happypath)
                        # pass的应用场景
                        Excel_text.enter_PASS('B14')
                        Excel_text.enter_Time('C14', start_time)
                        test_status = 1
                        break
                elif case_flag in item:
                    print(0)
                    crash()
                    search_status = 1
def SOC120(case_flag,butt):
    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()
    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)
    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)
    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)
    # 初始化输出列表
    output_list = []

    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    # Case33 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 6000:
            # fail的应用场景
            print(unhappypath)
            Excel_text.enter_FAIL('B27')
            Excel_text.enter_Time('C27', start_time)
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)

        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                if case_flag in item:
                    print(happypath)
                    # pass的应用场景
                    Excel_text.enter_PASS('B27')
                    Excel_text.enter_Time('C27', start_time)
                    test_status = 1
                    canalyzer().stop_Measurement()  # 关闭CAN工程
                    break
def SOCInhibit(case_flag,butt):
    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()

    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)

    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)

    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)

    # 初始化输出列表
    output_list = []

    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    # Case33 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 6000:
            # fail的应用场景
            print(unhappypath)
            Excel_text.enter_FAIL('B28')
            Excel_text.enter_Time('C28', start_time)
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)

        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                if case_flag in item:
                    print(happypath)
                    # pass的应用场景
                    Excel_text.enter_PASS('B28')
                    Excel_text.enter_Time('C28', start_time)
                    test_status = 1
                    canalyzer().stop_Measurement()
                    break
def DirectConfig1(case_flag,butt):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    # sheet1
    ws1 = wb.active
    ws1.title = "PDM"
    ws1.cell(row=1, column=1, value='用例:')
    ws1.cell(row=1, column=2, value='结果:')

    # 保存并生成文件
    wb.save('result.xlsx')

    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()

    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)

    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)

    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)

    # 初始化输出列表
    output_list = []

    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    # Case33 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 6000:
            # fail的应用场景
            print(unhappypath)
            wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
            ws = wb.active
            # 获取最后一行的行号
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value=butt)
            ws.cell(row=last_row, column=2, value=unhappypath)
            wb.save('result.xlsx')
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)

        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                if case_flag in item:
                    print(happypath)
                    # pass的应用场景
                    wb = openpyxl.load_workbook(
                        r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
                    ws = wb.active
                    # 获取最后一行的行号
                    last_row = ws.max_row + 1
                    ws.cell(row=last_row, column=1, value=butt)
                    ws.cell(row=last_row, column=2, value=happypath)
                    wb.save(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
                    test_status = 1
                    break

def Load_Shed(case_flag):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    # sheet1
    ws1 = wb.active
    ws1.title = "PDM"
    ws1.cell(row=1, column=1, value='用例:')
    ws1.cell(row=1, column=2, value='结果:')

    # 保存并生成文件
    wb.save('result.xlsx')

    # 创建SSH客户端
    ssh_client = paramiko.SSHClient()

    # 允许自动添加主机密钥
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 导入密钥
    private_key_path = 'id_ecg'
    private_key = paramiko.RSAKey.from_private_key_file(private_key_path)

    # 连接到远程服务器
    ssh_client.connect('10.2.0.1', username='root', pkey=private_key)

    # 执行命令并将输出重定向到文件
    command = 'fdpcat'
    stdin, stdout, stderr = ssh_client.exec_command(command)

    # 初始化输出列表
    output_list = []

    # 每2秒读取一次输入内容并打印前2秒内的所有内容
    initial_time = time.time()
    start_time = time.time()
    test_status = 0
    # Case33 结果写入报告中
    happypath = "PASS"
    unhappypath = "FAIL"
    while True:
        if test_status == 1:
            break
        if time.time() - initial_time > 6000:
            # fail的应用场景
            print(unhappypath)
            wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
            ws = wb.active
            # 获取最后一行的行号
            last_row = ws.max_row
            ws.cell(row=last_row, column=2, value=unhappypath)
            wb.save('result.xlsx')
            break
        if stdout.channel.recv_ready():
            output = stdout.channel.recv(1024).decode('latin-1')
            # print(output)
            output_list.append(output)
        else:
            time.sleep(0.1)

        # 判断是否达到2秒
        if time.time() - start_time >= 2:
            # print("前2秒内的输出内容:")
            for item in output_list:
                if case_flag in item:
                    print(happypath)
                    # pass的应用场景
                    wb = openpyxl.load_workbook(
                        r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
                    ws = wb.active
                    # 获取最后一行的行号
                    last_row = ws.max_row
                    ws.cell(row=last_row, column=2, value=happypath)
                    wb.save(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
                    test_status = 1
                    break



    # print("------------------------------------")
    # 重置计时器和输出列表
    start_time = time.time()
    output_list = []
# 读取输出内容
    ssh_client.close()



