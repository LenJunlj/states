import tkinter as tk
from test import open_Fenix
from canoe_open import canalyzer
from sshtestcase11_120 import crash120s
from sshtestcase11_120 import crash1023
from sshtestcase11_120 import SOC120
from sshtestcase11_120 import *
from tkinter import ttk
import time
from openpyxl import load_workbook
import openpyxl
from pycan import crash
from Excel_text import *

# 定义窗口
root=tk.Tk()
# print(type(root))
root.title('OTAgui')
root.geometry('800x600')

# 定义标签位置
Label_1 = tk.Label(root)
Label_1[ 'justify' ] = 'left'
Label_1[ 'text' ] = 'Ota_ChromeDriver:'
Label_1.place(x=5,y=10,height=20)

# 定义文本框内容
Text_1 = tk.Entry(root)
Text_1.insert(0,'C:\chromedriver_win32\chromedriver.exe')        # 文本框的起始位置插入默认值
Text_1.place(x=140,y=10,height=20,width=200)

Label_2 = tk.Label(root)
Label_2[ 'justify' ] = 'left'
Label_2[ 'text' ] = 'Ota_FinexUserId:'
Label_2.place(x=5,y=40,height=20)

Text_2 = tk.Entry(root)
Text_2.insert(0,'YYU66@ford.com')
Text_2.place(x=140,y=40,height=20,width=200)

Label_3 = tk.Label(root)
Label_3[ 'justify' ] = 'left'
Label_3[ 'text' ] = 'Ota_FinexEnvironment:'
Label_3.place(x=5,y=70,height=20)

Text_3 = tk.Entry(root)
Text_3.insert(0,'production')
Text_3.place(x=140,y=70,height=20,width=200)

Label_4 = tk.Label(root)
Label_4[ 'justify' ] = 'left'
Label_4[ 'text' ] = 'Ota_RolloutDescription:'
Label_4.place(x=5,y=100,height=20)

Text_4 = tk.Entry(root)
Text_4.insert(0,'GSM')
Text_4.place(x=140,y=100,height=20,width=200)

Label_5 = tk.Label(root)
Label_5[ 'justify' ] = 'left'
Label_5[ 'text' ] = 'Ota_VehicleSet:'
Label_5.place(x=5,y=130,height=20)

Text_5 = tk.Entry(root)
Text_5.insert(0,'CDX707_Prod_TT')
Text_5.place(x=140,y=130,height=20,width=200)

Label_6 = tk.Label(root)
Label_6[ 'justify' ] = 'left'
Label_6[ 'text' ] = 'Ota_VehicleVin:'
Label_6.place(x=5,y=160,height=20)

Text_6 = tk.Entry(root)
Text_6.insert(0,'279')
Text_6.place(x=140,y=160,height=20,width=200)

Label_7 = tk.Label(root)
Label_7[ 'justify' ] = 'left'
Label_7[ 'text' ] = 'Ota_SoftwareSet:'
Label_7.place(x=5,y=190,height=20)

Text_7 = tk.Entry(root)
Text_7[ 'text' ] = ''
Text_7.place(x=140,y=190,height=20,width=200)

Label_8 = tk.Label(root)
Label_8[ 'justify' ] = 'left'
Label_8[ 'text' ] = 'Ota_Install_Strategy:'
Label_8.place(x=5,y=220,height=20)

Text_8 = tk.Entry(root)
Text_8.insert(0,'Null')
Text_8.place(x=140,y=220,height=20,width=200)

Label_9 = tk.Label(root)
Label_9[ 'justify' ] = 'left'
Label_9[ 'text' ] = 'Ota_Inhibit_Required:'
Label_9.place(x=5,y=250,height=20)

Text_9 = tk.Entry(root)
Text_9.insert(0,'Y')
Text_9.place(x=140,y=250,height=20,width=200)

Label_10 = tk.Label(root)
Label_10[ 'justify' ] = 'left'
Label_10[ 'text' ] = 'Ota_UploadJsonDirect:'
Label_10.place(x=5,y=280,height=20)

Text_10 = tk.Entry(root)
Text_10.insert(0,'yes')
Text_10.place(x=140,y=280,height=20,width=200)


Label_11 = tk.Label(root)
Label_11[ 'justify' ] = 'left'
Label_11[ 'text' ] = 'Ota_SWStype:'
Label_11.place(x=5,y=310,height=20)

Text_11 = tk.Entry(root)
Text_11.insert(0,'单件改DC')
Text_11.place(x=140,y=310,height=20,width=200)

Label_12 = tk.Label(root)
Label_12[ 'justify' ] = 'left'
Label_12[ 'text' ] = 'Ota_JsonfilePath:'
Label_12.place(x=5,y=340,height=20)

Text_12 = tk.Entry(root)
Text_12.insert(0,r'C:\Users\YYU66\Desktop\MMOTA\Config\schema_dc_sample_json.json')
Text_12.place(x=140,y=340,height=20,width=200)


Label_13 = tk.Label(root)
Label_13[ 'justify' ] = 'left'
Label_13[ 'text' ] = 'Ota_JsonfilePath1:'
Label_13.place(x=5,y=370,height=20)

Text_13 = tk.Entry(root)
Text_13.insert(0,r'C:\Users\YYU66\Desktop\MMOTA\Config\schema_dc_sample_json_New.json')
Text_13.place(x=140,y=370,height=20,width=200)


Label_14 = tk.Label(root)
Label_14[ 'justify' ] = 'left'
Label_14[ 'text' ] = 'Ota_ExpirationHours:'
Label_14.place(x=5,y=400,height=20)

Text_14 = tk.Entry(root)
Text_14[ 'text' ] = ''
Text_14.place(x=140,y=400,height=20,width=200)


Label_15 = tk.Label(root)
Label_15[ 'justify' ] = 'left'
Label_15[ 'text' ] = 'Ota_UserConsentLevel:'
Label_15.place(x=5,y=430,height=20)


Text_15 = tk.Entry(root)
Text_15.insert(0,'1')
Text_15.place(x=140,y=430,height=20,width=200)


def start():   # 获取开始按钮相关联的文本框的内容
    open_Fenix.log_in(Text_1.get(),Text_2.get(),Text_3.get(),Text_4.get(),Text_5.get(),Text_6.get(),Text_7.get(),Text_8.get(),Text_9.get(),Text_10.get(),Text_11.get(),Text_12.get(),Text_13.get(),Text_14.get(),Text_15.get())
    # print(Text_1.get())
    print("文本框内容")


Label_16 = tk.Label(root)
Label_16[ 'justify' ] = 'left'
Label_16[ 'text' ] = 'Configruation:'
Label_16.place(x=5,y=460,height=20)

# 创建Combobox下拉框小部件
combo = ttk.Combobox(root)
combo['values'] = ('Case1', 'Case3', 'Case9', 'Case11_1023','Case11_120s','Case32','Case33','Case34','Case52','Case51')
combo.place(x=140,y=460)



# def compared():
#     item = combo.get()
#     wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
#     ws = wb.active
#     # 获取最后一行的行号
#     last_row = ws.max_row + 1
#     ws.cell(row=last_row, column=1, value=item)
#     wb.save('result.xlsx')
#     print(item)
#     elif item == "Case32":
#         itempath = r"C:\Users\YYU66\Desktop\MMOTA\Config\Configuration_12.0Lin\Configuration_12.0LIN_Load_shed(case32)\Configuration1.cfg"
#         canalyzer().open_cfg(itempath)
#         time.sleep(5)
#         canalyzer().start_Measurement()
#         time.sleep(10)
#         canalyzer().stop_Measurement()
#     elif item =="Case9":
#         itempath = r"C:\Users\YYU66\Desktop\MMOTA\Config\Configuration_12.0Lin\Configuration_12.0LIN_Load_shed(1023)\Configuration1.cfg"
#     print('CaseXX')

def start_button():
    if combo.get() == 'Case11_120s':
        itempath = r"C:\Users\YYU66\Desktop\MMOTA\Config\Configuration_12.0Lin\Configuration_12.0LIN_Crash_120\Configuration1_crash_120s.cfg"
        canalyzer().open_cfg(itempath)
        time.sleep(10)
        canalyzer().start_Measurement()
        time.sleep(20)
        case_f = "OTAM_S1011"
        set=Text_7.get()
        Excel_text.enter_Software('D15',set)
        butt=combo.get()
        crash120s(case_f,butt)

    elif combo.get() == 'Case11_1023':
        case_f = 'OTAM_S1023'
        butt = combo.get()
        crash1023(case_f,butt)
        print(1)

    elif combo.get() =='Case33':
        itempath =r"C:\Users\YYU66\Desktop\MMOTA\Config\Configuration_12.0Lin\Configuration_12.0LIN_SOC_120\Configuration_12.0LIN_SOC.cfg"
        canalyzer().open_cfg(itempath)
        time.sleep(10)
        canalyzer().start_Measurement()
        time.sleep(20)
        case_f ='OTAM_S1010'
        butt = combo.get()
        SOC120(case_f,butt)

    elif combo.get() =='Case34':
        itempath =r"C:\Users\YYU66\Desktop\MMOTA\Config\Configuration_12.0Lin\Configuration_12.0LIN_SOC_Inhibit\SOC_Inhibit.cfg"
        canalyzer().open_cfg(itempath)
        time.sleep(10)
        canalyzer().start_Measurement()
        time.sleep(20)
        case_f ='OTAM_S1010'
        butt=combo.get()
        SOCInhibit(case_f,butt)

    elif combo.get() =='Case1':
        case_f ='OTAM_S1010'
        butt = combo.get()
        DirectConfig1(case_f,butt)
        print('xxx')

def call_multiple_functions():
    start()
    start_button()

# 定义开始按钮
Start_button = tk.Button(root,command = call_multiple_functions)
Start_button['text'] = 'Start'
Start_button.place(x=150,y=520)

# 运行主循环
root.mainloop()
