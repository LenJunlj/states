import openpyxl
from openpyxl import workbook,Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles import Font
from fpdf import FPDF

def create_excle():
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    ws1.title = "PDM"
    ##设置列宽
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 30

################创建标题边框#################
    thin_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    for row in ws1['A2:B4']:
        for cell in row:
            cell.border = thin_border
    fill = PatternFill(fill_type='solid', fgColor='87CEEB')
    for row in ws1['A2:A4']:
        for cell in row:
            cell.fill = fill
    data = {
        'A2': 'CarLine',
        'A3': 'Function',
        'A4': 'Capture',
        'B2': 'CDX707C',
        'B3': 'MMOTA'
    }
    font = Font(name='Arial', size=12, bold=True, italic=True)
    for cell, value in data.items():
        ws1[cell] = value
        ws1[cell].font = font

################创建CASE边框######################3

    thin_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    for row in ws1['A6:D100']:
        for cell in row:
            cell.border = thin_border

    fill = PatternFill(fill_type='solid', fgColor='87CEEB')
    for row in ws1['A6:D6']:
        for cell in row:
            cell.fill = fill
    data = {
        'A6': {'value': 'Test Case ID'},
        'B6': {'value': 'Verdict(PASS/FAIL)'},
        'C6': {'value': 'Test Time'},
        'D6': {'value': 'Software Version'}
    }
    font = Font(name='Arial', size=12, bold=True, italic=True)
    for cell, cell_data in data.items():
        value = cell_data['value']
        ws1[cell] = value
        ws1[cell].font = font

    data = ['CASE 1', 'CASE 3', 'CASE 6', 'CASE 7', 'CASE 8','CASE 9','CASE 10','CASE 11_1023','CASE 11_120','CASE 16','CASE 18','CASE 19','CASE 21',
            'CASE 22','CASE 23','CASE 24','CASE 25','CASE 28','CASE 30','CASE 32','CASE 33','CASE 34','CASE 38','CASE 39','CASE 40','CASE 41','CASE 42',
            'CASE 49','CASE 52','CASE 53','CASE 54','CASE 55','CASE 56','CASE 57','CASE 58','CASE 59','CASE 61','CASE 63','CASE 65','CASE 66','CASE 67','CASE 68']
    font = Font(name='Arial', size=12, bold=False, italic=True)
    fill = PatternFill(start_color="D3D3FA", end_color="D3D3FA", fill_type="solid")
    start_row=7
    column_index=1
    for i, value in enumerate(data, start=start_row):
        cell = ws1.cell(row=i, column=column_index)
        cell.value = value
        cell.font = font
        cell.fill = fill
    # ws1.cell(row=1, column=1, value='Test case ID:')
    # ws1.cell(row=1, column=2, value='Result:')
    wb.save('result.xlsx')

def enter_PASS(position):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    font = Font(name='Arial', size=12, bold=False, italic=True)
    cell=ws1[position]
    cell.value='PASS'
    cell.fill = fill
    cell.font=font
    wb.save('result.xlsx')

def enter_FAIL(position):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font = Font(name='Arial', size=12, bold=False, italic=True)
    cell=ws1[position]
    cell.value='FAIL'
    cell.fill = fill
    cell.font=font
    wb.save('result.xlsx')

def enter_NA(position):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    font = Font(name='Arial', size=12, bold=False, italic=True)
    cell=ws1[position]
    cell.value='NA'
    cell.fill = fill
    cell.font=font
    wb.save('result.xlsx')


def enter_Time(position,start_time):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    font = Font(name='Arial', size=12, bold=False, italic=True)
    cell=ws1[position]
    cell.value=start_time
    cell.fill = fill
    cell.font=font
    wb.save('result.xlsx')

def enter_Software(position,set):
    wb = workbook.Workbook()
    wb = openpyxl.load_workbook(r'C:\Users\YYU66\Desktop\MMOTA\Project\pythonProject\result.xlsx')
    ws1 = wb.active
    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    font = Font(name='Arial', size=12, bold=False, italic=True)
    cell=ws1[position]
    cell.value=set
    cell.fill = fill
    cell.font=font
    wb.save('result.xlsx')

def convert_excel_to_pdf(excel_file, pdf_file):
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.worksheets[0]
    pdf = FPDF()
    for row in worksheet.iter_rows():
        for cell in row:
            pdf.cell(40, 10, str(cell.value))
        pdf.ln()
    pdf.output(pdf_file)

# create_excle()
# enter_PASS('B7')
# enter_FAIL('B8')
# enter_NA('B9')
# enter_Time('C7','2023.5.31.14.20')
# enter_Software('D7','CD707C_DSMC_7C1_AC_ME')